#!/usr/bin/env python3
"""
aggregate_comments.py — Gộp comment/RFI từ nhiều CSV thành 1 register Excel.
                        Consolidate comments/RFIs from many CSVs into one xlsx.

Dùng openpyxl + csv (stdlib). / Requires openpyxl.

Usage:
    python aggregate_comments.py a.csv b.csv -o register.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:  # pragma: no cover
    print("ERROR: cần openpyxl — chạy: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

# Cột chuẩn của register / canonical columns
CANON = ["id", "subject", "comment", "author", "discipline", "status", "location", "date"]

# Ánh xạ tên cột nguồn -> cột chuẩn (đã hạ chữ thường) / source-header aliases
COLUMN_ALIASES = {
    "id": "id", "no": "id", "number": "id", "issue id": "id", "markup id": "id",
    "subject": "subject", "title": "subject", "topic": "subject",
    "comment": "comment", "comments": "comment", "description": "comment",
    "contents": "comment", "details": "comment", "text": "comment", "notes": "comment",
    "author": "author", "created by": "author", "assignee": "author", "reviewer": "author",
    "discipline": "discipline", "trade": "discipline", "category": "discipline", "type": "discipline",
    "status": "status", "state": "status",
    "location": "location", "sheet": "location", "page": "location", "space": "location", "grid": "location",
    "date": "date", "created at": "date", "created": "date", "due date": "date",
}


def normalize_row(row: dict, source: str) -> dict:
    out = {c: "" for c in CANON}
    for raw_key, value in row.items():
        if raw_key is None:
            continue
        canon = COLUMN_ALIASES.get(raw_key.strip().lower())
        if canon and not out[canon]:
            out[canon] = (value or "").strip()
    out["source"] = source
    return out


def read_csv(path: Path) -> list[dict]:
    rows = []
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for r in reader:
            rows.append(normalize_row(r, path.name))
    return rows


def dedup(rows: list[dict]) -> list[dict]:
    """Gộp trùng theo (subject, comment, location). / merge duplicates."""
    seen: dict[tuple, dict] = {}
    for r in rows:
        key = (r["subject"].lower(), r["comment"].lower(), r["location"].lower())
        if key in seen and any(key):
            existing = seen[key]
            srcs = set(existing["source"].split("; ")) | {r["source"]}
            existing["source"] = "; ".join(sorted(srcs))
        else:
            seen[key] = r
    return list(seen.values())


def build_workbook(rows: list[dict], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Register"
    headers = CANON + ["source"]
    ws.append([h.capitalize() for h in headers])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for col, h in enumerate(headers, 1):
        width = max(12, min(50, max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows]) + 2))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    ws.freeze_panes = "A2"

    # Summary sheet
    s = wb.create_sheet("Summary")
    s.append(["Tổng comment / Total", len(rows)])
    s.append([])
    s.append(["Theo trạng thái / By status", ""])
    for status, n in Counter(r["status"] or "(blank)" for r in rows).most_common():
        s.append([status, n])
    s.append([])
    s.append(["Theo bộ môn / By discipline", ""])
    for disc, n in Counter(r["discipline"] or "(blank)" for r in rows).most_common():
        s.append([disc, n])
    for cell in (s["A1"], s["A3"], s["A"][s.max_row - 1] if False else s["A3"]):
        cell.font = Font(bold=True)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="Aggregate comments from CSVs into xlsx.")
    ap.add_argument("csvs", nargs="+")
    ap.add_argument("-o", "--out", required=True)
    args = ap.parse_args(argv)

    all_rows: list[dict] = []
    for c in args.csvs:
        p = Path(c)
        if not p.is_file():
            print(f"ERROR: không tìm thấy: {c}", file=sys.stderr)
            return 2
        all_rows.extend(read_csv(p))

    before = len(all_rows)
    merged = dedup(all_rows)
    build_workbook(merged, Path(args.out))

    print(f"Đã đọc {before} comment từ {len(args.csvs)} file.")
    print(f"Sau khử trùng lặp: {len(merged)} comment.")
    print(f"Đã ghi register: {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
